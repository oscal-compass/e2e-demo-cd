# Copyright (c) 2025 The OSCAL Compass Authors. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limit
"""OSCAL transformation tasks."""
import pathlib
from typing import Dict, Iterator, List

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class SheetHelper:
    """SheetHelper."""

    def __init__(self, wb: Workbook, sn: str) -> None:
        """Initialize."""
        self.wb = wb
        self.sn = sn
        self.ws = self.wb[self.sn]

    def get_sn(self) -> int:
        """Get sheet name."""
        return self.sn

    def get_max_col(self) -> int:
        """Get max column."""
        return self.ws.max_column

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 4
        while row <= self.ws.max_row:
            yield row
            row += 1

    def get_cell_value(self, row: int, col: int) -> str:
        """Get cell value for given row and column name."""
        cell = self.ws.cell(row, col)
        return cell.value

    def put_cell_value(self, row: int, col: int, value: str) -> None:
        """Get cell value for given row and column name."""
        cell = self.ws.cell(row, col)
        cell.value = value

    @staticmethod
    def get_sheetname_prefixes() -> List[str]:
        """Get sheetnames prefixes."""
        rval = ['Level 1', 'Level 2']
        return rval

    @staticmethod
    def get_sheetname() -> str:
        """Get sheetname output."""
        rval = 'Combined Profiles'
        return rval


class ColHelper:
    """Col Helper."""

    @staticmethod
    def get_cis_control() -> int:
        """Get cis control col no."""
        return 3

    @staticmethod
    def get_nist_control() -> int:
        """Get nist control col no."""
        return 12


class CisToNistMappingHelper():
    """CisToNistMappingHelper."""

    def __init__(self, ipath: pathlib.Path) -> None:
        """Initialize."""
        self.ipath = ipath
        self.sn = 'All CIS Controls & Safeguards'
        self.wb = load_workbook(ipath)
        self.ws = self.wb[self.sn]
        self.sheet_helper = SheetHelper(self.wb, self.sn)

    def get_map(self, include_list: List[str]) -> Dict[str, List[str]]:
        """Get map."""
        rval = {}
        sheet_helper = self.sheet_helper
        col_no_cis_control = ColHelper.get_cis_control()
        col_no_nist_control = ColHelper.get_nist_control()
        for row in sheet_helper.row_generator():
            cis_control = sheet_helper.get_cell_value(row, col_no_cis_control)
            nist_control = sheet_helper.get_cell_value(row, col_no_nist_control)
            if not cis_control:
                continue
            cis_id = f'cisc-{cis_control}'
            if cis_id not in include_list:
                continue
            if cis_id not in rval.keys():
                rval[cis_id] = []
            if nist_control not in rval[cis_id]:
                rval[cis_id].append(nist_control)
        return rval

    def get_map_reverse(self, include_list: List[str]) -> Dict[str, List[str]]:
        """Get map reverse."""
        rmap = {}
        imap = self.get_map(include_list)
        for key in imap.keys():
            value = imap[key]
            for nist in value:
                if nist not in rmap.keys():
                    rmap[nist] = []
                rmap[nist].append(key)
        return rmap
